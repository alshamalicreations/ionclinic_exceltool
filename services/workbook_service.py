"""
workbook_service.py

Responsible for:
- Loading Excel workbooks
- Saving Excel workbooks
- Listing worksheet names
- Providing access to worksheets

This service is the ONLY place in the application
that should directly interact with openpyxl.
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class WorkbookService:
    """
    Handles loading and saving Excel workbooks.
    """

    def load(self, file_path: str | Path) -> Workbook:
        """
        Load an Excel workbook.

        Args:
            file_path: Path to the workbook.

        Returns:
            Loaded Workbook object.
        """
        return load_workbook(file_path)

    def save(self, workbook: Workbook, output_path: str | Path) -> None:
        """
        Save an Excel workbook.

        Args:
            workbook: Workbook object.
            output_path: Destination path.
        """
        workbook.save(output_path)

    def get_sheet_names(self, workbook: Workbook) -> list[str]:
        """
        Return all worksheet names.
        """
        return workbook.sheetnames

    def sheet_exists(self, workbook: Workbook, sheet_name: str) -> bool:
        """
        Check whether a worksheet exists.
        """
        return sheet_name in workbook.sheetnames